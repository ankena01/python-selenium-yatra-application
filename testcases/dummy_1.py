# import pandas as pd
# df = pd.read_excel(r"C:\temp\TestFrameWorkDemo\testdata\testdata.xlsx",sheet_name="Sheet1")
# #print(df)
# row_count = len(df)
# data_list=[]

# for i in range(0, row_count):
#     row = [] 
#     for x in df.iloc[i]:
#         row.append(x)
#     data_list.append(row)
# print(data_list)

from openpyxl import Workbook, load_workbook

def read_data_from_excel(file_name , sheet):
    datalist = []
    wb = load_workbook(filename=file_name)
    sh = wb[sheet]
    row_ct = sh.max_row
    col_ct = sh.max_column

    for i in range(2,row_ct+1):

        row = []
        for j in range(1,col_ct+1):
            row.append(sh.cell(row=i , column=j).value)
        datalist.append(row)
    return datalist

result = read_data_from_excel("C:\\temp\\TestFrameWorkDemo\\testdata\\testdata.xlsx","Sheet1")
print(result)

import os
import time

@pytest.hookimpl(hookwrapper=True)
def pytest_runtest_makereport(item, call):
    pytest_html = item.config.pluginmanager.getplugin("html")
    outcome = yield
    report = outcome.get_result()
    extra = getattr(report, "extra", [])
    if report.when == "call":
        # always add url to report
        extra.append(pytest_html.extras.url("http://www.example.com/"))
        xfail = hasattr(report, "wasxfail")
        if (report.skipped and xfail) or (report.failed and not xfail):
            # only add additional html on failure
            report_directory = os.path.dirname(item.config.option.htmlpath)
            file_name = str(int(round(time.time()*1000)))+".png"
            file_name = report.nodeid.replace("::","_") + ".png"
            destinationFile = os.path.join(report_directory , file_name)
            self.driver.save_screenshot(destinationFile)
            if file_name:
                html='<div><img src="%s" alt="screenshot" style="width:300px;height=200px"' \
                    'onclick="window.open(this.src)" align="right"/></div>'%file_name
            extra.append(pytest_html.extras.html(html))
        report.extra = extra

def pytest_html_report_title(report):
    report.title = "Test Automation FrameWork Report"


            extra.append(pytest_html.extras.html("<div>Additional HTML</div>"))
        report.extra = extra